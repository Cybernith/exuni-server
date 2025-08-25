# imports لازم
import os
import re
import zipfile
import xml.etree.ElementTree as ET
import traceback

from django.core.files.base import ContentFile
from django.db import transaction
from openpyxl import load_workbook

# --- helper ها (تقریبا شبیه کدی که قبلا دادم) ---
from file_handler.models import ExtractedImage


XML_NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'rels_pkg': 'http://schemas.openxmlformats.org/package/2006/relationships'
}


def _parse_cellimages_xml(zip_file):
    id_to_rid = {}
    path = "xl/cellimages.xml"
    if path not in zip_file.namelist():
        return id_to_rid
    try:
        xml = zip_file.read(path)
        ns = {
            'xdr': "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            'r': "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            'etc': "http://www.wps.cn/officeDocument/2017/etCustomData",
            'a': "http://schemas.openxmlformats.org/drawingml/2006/main",
        }
        root = ET.fromstring(xml)

        # پیدا کردن همه عناصر xdr:pic
        for pic in root.findall(".//xdr:pic", ns):
            # استخراج cNvPr برای گرفتن نام (img_id)
            cNvPr = pic.find("xdr:nvPicPr/xdr:cNvPr", ns)
            if cNvPr is not None:
                img_id = cNvPr.attrib.get("name")
                if img_id:
                    # استخراج blip برای گرفتن rid
                    blip = pic.find("xdr:blipFill/a:blip", ns)
                    if blip is not None:
                        rid = blip.attrib.get(f"{{{ns['r']}}}embed") or blip.attrib.get("r:embed")
                        if rid:
                            id_to_rid[img_id] = rid
    except Exception:
        print("ERROR parsing cellimages.xml:", traceback.format_exc())
    return id_to_rid

def _parse_cellimages_rels_et(zip_file):
    """
    برمی‌گرداند: rels_map => { 'rIdN' : 'image84.png' }
    با استفاده از ElementTree روی xl/_rels/cellimages.xml.rels یا فایل مشابه داخل xl/_rels/
    """
    rels_map = {}
    # مسیر استاندارد
    rels_path = "xl/_rels/cellimages.xml.rels"
    if rels_path not in zip_file.namelist():
        # گاهی نام فایل rels در لیست کمی متفاوت است؛ جستجوی هر فایل در xl/_rels که نامش cellimages در آن باشد
        for name in zip_file.namelist():
            if name.startswith("xl/_rels/") and 'cellimages' in name and name.endswith('.rels'):
                rels_path = name
                break
        else:
            return rels_map

    try:
        raw = zip_file.read(rels_path)
        root = ET.fromstring(raw)
        # Relationship elements ممکنه در namespace باشند؛ اما child iteration کار می‌دهد
        for rel in root:
            # هر عنصر Relationship را بررسی کن
            rid = rel.attrib.get('Id') or rel.attrib.get('ID') or rel.attrib.get('{http://schemas.openxmlformats.org/package/2006/relationships}Id')
            target = rel.attrib.get('Target') or rel.attrib.get('target')
            if rid and target:
                # target ممکنه "../media/image84.png" یا "media/image84.png"
                rels_map[rid] = os.path.basename(target)
    except Exception:
        print("ERROR parsing rels ET:", traceback.format_exc())
    return rels_map

def _extract_dispimg_ids_from_workbook(xlsx_path, sheet_name=None):
    id_to_row = {}
    try:
        wb = load_workbook(filename=xlsx_path, data_only=False)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        pattern = re.compile(r'DISPIMG\("([^"]+)"', re.IGNORECASE)
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if not val:
                    continue
                m = pattern.search(str(val))
                if m:
                    img_id = m.group(1)
                    id_to_row[img_id] = cell.row - 1
    except Exception:
        print("ERROR reading workbook formulas:", traceback.format_exc())
    return id_to_row

# ---------- تابع اصلی مقاوم و با لاگ ----------
def extract_images_from_excel_and_map_rows_debug(xlsx_path, instance, commit_to_db=True):
    """
    Returns: created_images, row_map, errors
    نسخه debug-friendly و امن برای position
    """
    created_images = []
    row_map = {}
    errors = []

    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # parse relationships
        rels_map = _parse_cellimages_rels_et(z)   # rId -> filename
        id_to_rid = _parse_cellimages_xml(z)      # imageID -> rId

        print("=== DEBUG: rels_map (rId -> filename) ===")
        print(rels_map)
        print("=== DEBUG: id_to_rid (imageID -> rId) ===")
        print(id_to_rid)

        # build id -> filename
        id_to_filename = {}
        for img_id, rId in id_to_rid.items():
            filename = rels_map.get(rId)
            if filename:
                id_to_filename[img_id] = filename

        # parse workbook formulas (DISPIMG IDs)
        id_to_row = _extract_dispimg_ids_from_workbook(xlsx_path)
        print("=== DEBUG: id_to_row (imageID -> row_index) ===")
        print(id_to_row)

        # build filename -> row_index
        filename_to_row = {}
        for img_id, filename in id_to_filename.items():
            key = os.path.basename(filename).strip()
            row = id_to_row.get(img_id)
            if row is not None:
                filename_to_row[key] = row
            else:
                print(f"WARNING: No row mapping for imageID={img_id}, filename={key}")

        print("=== DEBUG: filename_to_row (filename -> row_index) ===")
        print(filename_to_row)

        image_count = 0
        for name in z.namelist():
            if not name.startswith("xl/media/") or name.endswith("/"):
                continue

            try:
                img_bytes = z.read(name)
            except Exception as e:
                err = f"READ_MEDIA_ERROR {name}: {e}"
                print(err); print(traceback.format_exc()); errors.append(err)
                continue

            file_extension = name.split('.')[-1].lower()
            basename = os.path.basename(name)

            try:
                image_file = ContentFile(img_bytes, name=f"{instance.id}_{image_count}.{file_extension}")
            except Exception as e:
                err = f"CONTENTFILE_ERROR {name}: {e}"
                print(err); print(traceback.format_exc()); errors.append(err)
                continue

            row_index = filename_to_row.get(basename)
            if row_index is None:
                # fallback: عدد شمارنده image_count به عنوان position
                print(f"INFO: No row_index for {basename}, fallback to image_count={image_count}")
                row_index = image_count

            # position_value نهایی
            position_value = str(row_index)

            # check existence of excel_row field
            try:
                ExtractedImage._meta.get_field('excel_row')
                has_excel_row = True
            except Exception:
                has_excel_row = False

            if commit_to_db:
                try:
                    with transaction.atomic():
                        if has_excel_row:
                            img_instance = ExtractedImage.objects.create(
                                uploaded_file=instance,
                                image=image_file,
                                position=row_index,
                            )
                        else:
                            img_instance = ExtractedImage.objects.create(
                                uploaded_file=instance,
                                image=image_file,
                                position=position_value
                            )
                    created_images.append(img_instance)
                    row_map[row_index] = img_instance
                except Exception as e:
                    err = f"DB_CREATE_ERROR {name}: {e}"
                    print(err); print(traceback.format_exc()); errors.append(err)
                    continue
            else:
                created_images.append({
                    'temp_name': image_file.name,
                    'basename': basename,
                    'row_index': row_index,
                    'bytes_len': len(img_bytes),
                    'position_value': position_value
                })

            image_count += 1

    return created_images, row_map, errors
